# ========================== IMPORT STATEMENTS =============================================================
import sys

logfile = open('program_output.txt', 'w')
sys.stdout = logfile
sys.stderr = logfile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import eel
from sys import exit
import os
import eel.browsers
import ctypes
#============================= REMOVE DURING DEBUGGING ====================================================
import warnings
#warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
#workbook = load_workbook(filename="C:\\Personal\\Wayside-Installation-Tracker\\48012-Progress-Tracker.xlsx",  data_only=True)
current_directory = os.path.abspath(__file__).replace("\\dist\\WaysideInstallationTracker\\_internal\\WaysideInstallationTracker.py","")
file_path = current_directory + "\\48012-Progress-Tracker.xlsx"

workbook = load_workbook(filename=file_path,  data_only=True)

#========================================================================================================================================================================
#====================================================== CLASS DEFINITIONS, SEE GITHUB REPO FOR DETAILS ==================================================================
#========================================================================================================================================================================

#======================================================= GENERAL USE CLASSES/SUPERCLASSES =============================================================================

class BinProgress:
    def __init__(self, date, progress):
        self.date = date
        self.progress = progress

class Equipment:
    def __init__(self,stationing,track,location,activities,notes):
        self.stationing = stationing
        self.track = track
        self.location = location
        self.activities = activities
        self.notes = notes

class AXC(Equipment):
    def __init__(self,stationing,track,location,activities,notes,REFDWG):
        self.stationing = stationing
        self.track = track
        self.location = location
        self.activities = activities
        self.notes = notes
        self.REFDWG = REFDWG

    def getProgress(self):
        if (self.activities.ACInstall.progress is None) \
            or (self.activities.JBInstall.progress is None) \
            or (self.activities.LCInstall.progress is None) \
            or (self.activities.preOpTesting.progress is None):
            return -1
        return (self.activities.ACInstall.progress + self.activities.JBInstall.progress + self.activities.LCInstall.progress + self.activities.preOpTesting.progress)/4

class Signal(Equipment):
    def __init__(self,stationing,track,signalType,location,activities,notes):
        super().__init__(stationing,track,location,activities,notes)
        self.signalType = signalType
    
    def getProgress(self):
        if (self.activities.sigInstall.progress is None) \
            or (self.activities.JBInstall.progress is None) \
            or (self.activities.SMInstall.progress is None) \
            or (self.activities.LCInstall.progress is None) \
            or (self.activities.preOpTesting.progress is None):
            return -1
        return (self.activities.sigInstall.progress + self.activities.JBInstall.progress + self.activities.SMInstall.progress + self.activities.LCInstall.progress + self.activities.preOpTesting.progress)/5

class Switch(Equipment):
    def __init__(self,name,stationing,track,location,activities,notes):
        super().__init__(stationing,track,location,activities,notes)
        self.name = name

    def getProgress(self):
        if (self.activities.switchInstall.progress is None) \
            or (self.activities.JBInstall.progress is None) \
            or (self.activities.SCInstall.progress is None) \
            or (self.activities.LCInstall.progress is None) \
            or (self.activities.breakdownTesting.progress is None) \
            or (self.activities.preOpTesting.progress is None):
            return -1
        return (self.activities.switchInstall.progress + self.activities.JBInstall.progress + self.activities.SCInstall.progress + self.activities.LCInstall.progress + self.activities.breakdownTesting.progress + self.activities.preOpTesting.progress)/6

class WRU(Equipment):
    def __init__(self,stationing,track,location,activities,notes):
        super().__init__(stationing,track,location,activities,notes)
    
    def getProgress(self):
        if (self.activities.RUInstall.progress is None) \
            or (self.activities.JBInstall.progress is None) \
            or (self.activities.FBInstall.progress is None) \
            or (self.activities.antennaInstall.progress is None) \
            or (self.activities.antCableInstall.progress is None) \
            or (self.activities.splitterInstall.progress is None) \
            or (self.activities.FCSplice.progress is None) \
            or (self.activities.FTesting.progress is None) \
            or (self.activities.PTesting.progress is None):
            return -1
        return (self.activities.RUInstall.progress + self.activities.JBInstall.progress + self.activities.FBInstall.progress + self.activities.antennaInstall.progress + self.activities.antCableInstall.progress + self.activities.splitterInstall.progress + self.activities.FCSplice.progress + self.activities.FTesting.progress + self.activities.PTesting.progress)/6

class ZCase(Equipment):
    def __init__(self,stationing,track,location,activities,notes):
        super().__init__(stationing,track,location,activities,notes)

    def getProgress(self):
        if (self.activities.caseInstall.progress is None) \
            or (self.activities.cableConnect.progress is None) \
            or (self.activities.preOpTesting.progress is None):
            return -1
        return (self.activities.caseInstall.progress + self.activities.cableConnect.progress + self.activities.preOpTesting.progress)/3

class TOPB(Equipment):
    def __init__(self,stationing,track,location,activities,notes):
        super().__init__(stationing,track,location,activities,notes)

    def getProgress(self):
        if (self.activities.TOPBInstall.progress is None) \
            or (self.activities.cableConnect.progress is None) \
            or (self.activities.preOpTesting.progress is None):
            return -1
        return (self.activities.TOPBInstall.progress + self.activities.cableConnect.progress + self.activities.preOpTesting.progress)/3

#====================================================== CMRS CLASSES ==============================================================================================
class CableSpan:
    def __init__(self,start,end,track,location,type,activities,notes):
        self.start = start
        self.end = end
        self.track = track
        self.location = location
        self.type = type
        self.activities = activities
        self.notes = notes

    def __str__(self):
        return str(self.type).upper() + " " + str(self.start) + " to " + str(self.end)
    
    #returns the (unweighted) progress of the current cablespan as a decimal between 0 and 1. Returns -1 if totals are undefined
    def getProgress(self):
        if self.activities == None:
            raise ValueError("CableSpan does not have an activity defined")
        
        match type(self.activities).__name__:
            case "MessActivities":
                if (progress := self.activities.messSupports.getProgress()) is None \
                    or (progress := self.activities.messClamps.getProgress()) is None \
                    or (progress := self.activities.messWirePull.getProgress()) is None \
                    or (progress := self.activities.messWireTension.getProgress()) is None \
                    or (progress := self.activities.messCablesPulled.getProgress()) is None \
                    or (progress := self.activities.messStraps.getProgress()) is None:
                    return -1
                return (self.activities.messClamps.getProgress() + self.activities.messWirePull.getProgress() +
                        self.activities.messWireTension.getProgress() + self.activities.messCablesPulled.getProgress() +
                        self.activities.messStraps.getProgress()) / 6
            
            case "CMRS15Activities":
                if (progress := self.activities.colClamp.getProgress()) is None \
                        or (progress := self.activities.stationBrackets.getProgress()) is None \
                        or (progress := self.activities.grounding.getProgress()) is None \
                        or (progress := self.activities.obsBracket.getProgress()) is None \
                        or (progress := self.activities.cablesPulled.getProgress()) is None \
                        or (progress := self.activities.CMRSInstall15) is None:
                    return -1

                return (self.activities.CMRSInstall15.getProgress() + self.activities.colClamp.getProgress() + self.activities.stationBrackets.getProgress() + self.activities.grounding.getProgress() +
                        self.activities.obsBracket.getProgress() + self.activities.cablesPulled.getProgress()) / 6
            
            case "CMRS24Activities":
                if (progress := self.activities.colClamp.getProgress()) is None \
                        or (progress := self.activities.stationBrackets.getProgress()) is None \
                        or (progress := self.activities.grounding.getProgress()) is None \
                        or (progress := self.activities.obsBracket.getProgress()) is None \
                        or (progress := self.activities.cablesPulled.getProgress()) is None \
                        or (progress := self.activities.CMRSInstall24) is None:
                    return -1
                return (self.activities.CMRSInstall24.getProgress() + self.activities.colClamp.getProgress() + self.activities.stationBrackets.getProgress() + self.activities.grounding.getProgress() +
                        self.activities.obsBracket.getProgress() + self.activities.cablesPulled.getProgress()) / 5

            case "TrayActivities":
                if (progress := self.activities.trayBrackets.getProgress()) is None \
                        or (progress := self.activities.installingTray.getProgress()) is None \
                        or (progress := self.activities.coreDrilling.getProgress()) is None \
                        or (progress := self.activities.cablePull.getProgress()) is None:
                    return -1
                return (self.activities.trayBrackets.getProgress() + self.activities.installingTray.getProgress() + self.activities.coreDrilling.getProgress() +
                        self.activities.cablePull.getProgress()) / 4

            case _: #dump out 
                raise ValueError("Unknown activity type")
            
class MessActivities:
    def __init__(self, messSupports, messClamps, messWirePull, messWireTension, messCablesPulled,messStraps):
        self.messSupports = messSupports
        self.messClamps = messClamps
        self.messWirePull = messWirePull
        self.messWireTension = messWireTension
        self.messCablesPulled = messCablesPulled
        self.messStraps = messStraps

        
class CMRSActivities:
    def __init__(self, colClamp, stationBrackets, grounding, obsBracket, cablesPulled):
        self.colClamp = colClamp
        self.stationBrackets = stationBrackets
        self.grounding = grounding
        self.obsBracket = obsBracket
        self.cablesPulled = cablesPulled

class CMRS15Activities(CMRSActivities):
    def __init__(self, CMRSInstall15, colClamp, stationBrackets, grounding, obsBracket, cablesPulled):
        super().__init__(colClamp, stationBrackets, grounding, obsBracket, cablesPulled)
        self.CMRSInstall15 = CMRSInstall15

class CMRS24Activities(CMRSActivities):
    def __init__(self, CMRSInstall24, colClamp, stationBrackets, grounding, obsBracket, cablesPulled):
        super().__init__(colClamp, stationBrackets, grounding, obsBracket, cablesPulled)
        self.CMRSInstall24 = CMRSInstall24

class TrayActivities:
    def __init__(self, trayBrackets, installingTray, coreDrilling, cablePull):
        self.trayBrackets = trayBrackets
        self.installingTray = installingTray
        self.coreDrilling = coreDrilling
        self.cablePull = cablePull

class progress:
    def __init__(self, total, install):
        self.total = total
        self.install = install

    #returns progress as a decimal between 0 and 1. If self.install is undefined, automatically default to 0
    def getProgress(self):
        if self.total == None:
            return None
        if self.total > 0:
            if self.install == None:
                self.install = 0
            return self.install/self.total
        else:
            return None
            #raise ValueError("the \"total\" of this activity is 0")

#========================================================================== AXLE COUNTER ============================================================================

class AxleCounterActivities:
    def __init__(self,ACInstall, JBInstall, LCInstall, ECInstall, preOpTesting):
        self.ACInstall = ACInstall
        self.JBInstall = JBInstall
        self.LCInstall = LCInstall
        self.ECInstall = ECInstall
        self.preOpTesting = preOpTesting

#======================================================================== SIGNALS ===================================================================================

class SignalActivities:
    def __init__(self, sigInstall, JBInstall, SMInstall, LCInstall, breakdownTesting, preOpTesting):
        self.sigInstall = sigInstall
        self.JBInstall = JBInstall
        self.SMInstall = SMInstall
        self.LCInstall = LCInstall
        self.breakdownTesting = breakdownTesting
        self.preOpTesting = preOpTesting

# ==================================================================== SWITCH ======================================================================================

class SwitchActivities:
    def __init__(self, switchInstall, JBInstall, SCInstall, LCInstall, breakdownTesting, preOpTesting):
        self.switchInstall = switchInstall
        self.JBInstall = JBInstall
        self.SCInstall = SCInstall
        self.LCInstall = LCInstall
        self.breakdownTesting = breakdownTesting
        self.preOpTesting = preOpTesting

# ==================================================================== WRU ========================================================================================

class WRUActivities:
    def __init__(self, RUInstall, JBInstall, FBInstall, antennaInstall, antCableInstall, splitterInstall, FCSplice, FTesting, PTesting):
        self.RUInstall = RUInstall
        self.JBInstall = JBInstall
        self.FBInstall = FBInstall
        self.antennaInstall = antennaInstall
        self.antCableInstall = antCableInstall
        self.splitterInstall = splitterInstall
        self.FCSplice = FCSplice
        self.FTesting = FTesting
        self.PTesting = PTesting

# ==================================================================== ZCase =======================================================================================

class ZCaseActivities:
    def __init__(self, caseInstall, cableConnect, preOpTesting):
        self.caseInstall = caseInstall
        self.cableConnect = cableConnect
        self.preOpTesting = preOpTesting

# ==================================================================== TOPB ====================================================================================
    
class TOPBActivities:
    def __init__(self, TOPBInstall, cableConnect, preOpTesting):
        self.TOPBInstall = TOPBInstall
        self.cableConnect = cableConnect
        self.preOpTesting = preOpTesting

#============================================================ END CLASS DEFINITIONS ============================================================================

# takes a string and returns an int, if None is passed, None is returned
def toNum(intVal):
    if intVal == None:
        return None
    try:
        return float(intVal)
    except ValueError:
        return None

# reads a block of CMRS data, corresponding to one entry, and converts it to a CableSpan object. Returns a CableSpan object. 
def readCMRSBlock(startingRow):
    sheet = workbook["CMRS"] #Open CMRS workbook
    match sheet.cell(startingRow,5).value: #Read the entry on the "activity" column
        case "Steel Mess Supports": #This is the first entry for messenger supports
            MessengerSpan = CableSpan(toNum(sheet.cell(startingRow,1).value),toNum(sheet.cell(startingRow,2).value),str(sheet.cell(startingRow,3).value),str(sheet.cell(startingRow,4).value),"mess",None,None)
            
            sMSupp = progress(toNum(sheet.cell(startingRow,6).value),toNum(sheet.cell(startingRow,7).value))
            mClamps = progress(toNum(sheet.cell(startingRow+1,6).value),toNum(sheet.cell(startingRow+1,7).value))
            mWPull = progress(toNum(sheet.cell(startingRow+2,6).value),toNum(sheet.cell(startingRow+2,7).value))
            mWTension = progress(toNum(sheet.cell(startingRow+3,6).value),toNum(sheet.cell(startingRow+3,7).value))
            nCPull = progress(toNum(sheet.cell(startingRow+4,6).value),toNum(sheet.cell(startingRow+4,7).value))
            mStraps = progress(toNum(sheet.cell(startingRow+5,6).value),toNum(sheet.cell(startingRow+5,7).value))
            
            MessengerSpan.activities = MessActivities(sMSupp,mClamps,mWPull,mWTension,nCPull,mStraps)
            return MessengerSpan
        
        case "15\" CMRS Install":
            CMRS15Span = CableSpan(toNum(sheet.cell(startingRow,1).value),toNum(sheet.cell(startingRow,2).value),str(sheet.cell(startingRow,3).value),str(sheet.cell(startingRow,4).value),"15CMRS",None,None)
            
            cInstall = progress(toNum(sheet.cell(startingRow,6).value),toNum(sheet.cell(startingRow,7).value))
            cClamp = progress(toNum(sheet.cell(startingRow+1,6).value),toNum(sheet.cell(startingRow+1,7).value))
            sBracket = progress(toNum(sheet.cell(startingRow+2,6).value),toNum(sheet.cell(startingRow+2,7).value))
            grounding = progress(toNum(sheet.cell(startingRow+3,6).value),toNum(sheet.cell(startingRow+3,7).value))
            oBracket = progress(toNum(sheet.cell(startingRow+4,6).value),toNum(sheet.cell(startingRow+4,7).value))
            nCPull = progress(toNum(sheet.cell(startingRow+5,6).value),toNum(sheet.cell(startingRow+5,7).value))

            CMRS15Span.activities = CMRS15Activities(cInstall,cClamp,sBracket,grounding,oBracket,nCPull)
            return CMRS15Span
        
        case "24\" CMRS Install":
            CMRS24Span = CableSpan(toNum(sheet.cell(startingRow,1).value),toNum(sheet.cell(startingRow,2).value),str(sheet.cell(startingRow,3).value),str(sheet.cell(startingRow,4).value),"24CMRS",None,None)
            
            cInstall = progress(toNum(sheet.cell(startingRow,6).value),toNum(sheet.cell(startingRow,7).value))
            cClamp = progress(toNum(sheet.cell(startingRow+1,6).value),toNum(sheet.cell(startingRow+1,7).value))
            sBracket = progress(toNum(sheet.cell(startingRow+2,6).value),toNum(sheet.cell(startingRow+2,7).value))
            grounding = progress(toNum(sheet.cell(startingRow+3,6).value),toNum(sheet.cell(startingRow+3,7).value))
            oBracket = progress(toNum(sheet.cell(startingRow+4,6).value),toNum(sheet.cell(startingRow+4,7).value))
            nCPull = progress(toNum(sheet.cell(startingRow+5,6).value),toNum(sheet.cell(startingRow+5,7).value))

            CMRS24Span.activities = CMRS24Activities(cInstall,cClamp,sBracket,grounding,oBracket,nCPull)
            return CMRS24Span
        case "Cable Tray Brackets":
            TraySpan = CableSpan(toNum(sheet.cell(startingRow,1).value),toNum(sheet.cell(startingRow,2).value),str(sheet.cell(startingRow,3).value),str(sheet.cell(startingRow,4).value),"tray",None,None)

            ctBrackets = progress(toNum(sheet.cell(startingRow,6).value),toNum(sheet.cell(startingRow,7).value))
            iTray = progress(toNum(sheet.cell(startingRow+1,6).value),toNum(sheet.cell(startingRow+1,7).value))
            cDrilling = progress(toNum(sheet.cell(startingRow+2,6).value),toNum(sheet.cell(startingRow+2,7).value))
            cPull = progress(toNum(sheet.cell(startingRow+3,6).value),toNum(sheet.cell(startingRow+3,7).value))

            TraySpan.activities = TrayActivities(ctBrackets,iTray,cDrilling,cPull)
            return TraySpan
        case "Crossover (Arch Bar Installation)":
            return None
        case _:
            err_message = "ERR | " + str(sheet.cell(startingRow,5).value) + "is NOT recognized as a cablespan type"
            print(err_message)


CMRSObjectList = []

def readCMRSWorksheet():
    currRowIndex = 3
    while currRowIndex <= 1215:
        block1 = readCMRSBlock(currRowIndex)
        CMRSObjectList.append(block1)
        if block1 == None:
            currRowIndex += 1
        else:
            match block1.type:
                case "mess":
                    currRowIndex += 6
                case "15CMRS":
                    currRowIndex += 6
                case "24CMRS":
                    currRowIndex += 6
                case "tray":
                    currRowIndex += 4

def readAXCBlock(startingRow):
    sheet = workbook["AXLE COUNTER"]
    axcObj = AXC(toNum(sheet.cell(startingRow,1).value),toNum(sheet.cell(startingRow,2).value),str(sheet.cell(startingRow,3).value),None,None,None)

    AC = BinProgress(toNum(sheet.cell(startingRow,5).value),toNum(sheet.cell(startingRow,6).value))
    JB = BinProgress(toNum(sheet.cell(startingRow+1,5).value),toNum(sheet.cell(startingRow+1,6).value))
    LC = BinProgress(toNum(sheet.cell(startingRow+2,5).value),toNum(sheet.cell(startingRow+2,6).value))
    EC = BinProgress(toNum(sheet.cell(startingRow+3,5).value),toNum(sheet.cell(startingRow+3,6).value))
    POT = BinProgress(toNum(sheet.cell(startingRow+4,5).value),toNum(sheet.cell(startingRow+4,6).value))

    axcObj.activities = AxleCounterActivities(AC,JB,LC,EC,POT)
    return axcObj

AXCObjectList = []

def readAXCWorksheet():
    currRowIndex = 2
    while currRowIndex <= 387:
        AXCObjectList.append(readAXCBlock(currRowIndex))
        currRowIndex += 5

def readSignalBlock(startingRow):
    sheet = workbook["SIGNALS"]
    signalObj = Signal(toNum(sheet.cell(startingRow,1).value),str(sheet.cell(startingRow,2).value),str(sheet.cell(startingRow,3).value),str(sheet.cell(startingRow,4).value),None,None)

    SI = BinProgress(toNum(sheet.cell(startingRow,6).value),toNum(sheet.cell(startingRow,7).value))
    JB = BinProgress(toNum(sheet.cell(startingRow+1,6).value),toNum(sheet.cell(startingRow+1,7).value))
    SM = BinProgress(toNum(sheet.cell(startingRow+2,6).value),toNum(sheet.cell(startingRow+2,7).value))
    LC = BinProgress(toNum(sheet.cell(startingRow+3,6).value),toNum(sheet.cell(startingRow+3,7).value))
    BT = BinProgress(toNum(sheet.cell(startingRow+4,6).value),toNum(sheet.cell(startingRow+4,7).value))
    POT = BinProgress(toNum(sheet.cell(startingRow+5,6).value),toNum(sheet.cell(startingRow+5,7).value))
    
    signalObj.activities = SignalActivities(SI,JB,SM,LC,BT,POT)
    return signalObj

SignalObjectList = []

def readSignalWorksheet():
    currRowIndex = 2
    while currRowIndex <= 194:
        SignalObjectList.append(readSignalBlock(currRowIndex))
        currRowIndex += 6

def readSwitchBlock(startingRow):
    sheet = workbook["SWITCH"]
    switchObj = Switch(str(sheet.cell(startingRow,1).value),toNum(sheet.cell(startingRow,2).value),str(sheet.cell(startingRow,3).value),str(sheet.cell(startingRow,4).value),None,None)

    SI = BinProgress(toNum(sheet.cell(startingRow,6).value),toNum(sheet.cell(startingRow,7).value))
    JB = BinProgress(toNum(sheet.cell(startingRow+1,6).value),toNum(sheet.cell(startingRow+1,7).value))
    SC = BinProgress(toNum(sheet.cell(startingRow+2,6).value),toNum(sheet.cell(startingRow+2,7).value))
    LC = BinProgress(toNum(sheet.cell(startingRow+3,6).value),toNum(sheet.cell(startingRow+3,7).value))
    BT = BinProgress(toNum(sheet.cell(startingRow+4,6).value),toNum(sheet.cell(startingRow+4,7).value))
    POT = BinProgress(toNum(sheet.cell(startingRow+5,6).value),toNum(sheet.cell(startingRow+5,7).value))

    switchObj.activities = SwitchActivities(SI,JB,SC,LC,BT,POT)
    return switchObj

SwitchObjectList = []

def readSwitchWorksheet():
    currRowIndex = 2
    while currRowIndex <= 134:
        SwitchObjectList.append(readSwitchBlock(currRowIndex))
        currRowIndex += 6

def readWRUBlock(startingRow):
    sheet = workbook["WRU"]
    WRUObj = WRU(toNum(sheet.cell(startingRow,1).value),str(sheet.cell(startingRow,2).value),str(sheet.cell(startingRow,3).value),None,None)

    RU = BinProgress(toNum(sheet.cell(startingRow,5).value),toNum(sheet.cell(startingRow,6).value))
    JB = BinProgress(toNum(sheet.cell(startingRow+1,5).value),toNum(sheet.cell(startingRow+1,6).value))
    FB = BinProgress(toNum(sheet.cell(startingRow+2,5).value),toNum(sheet.cell(startingRow+2,6).value))
    ANT = BinProgress(toNum(sheet.cell(startingRow+3,5).value),toNum(sheet.cell(startingRow+3,6).value))
    AC = BinProgress(toNum(sheet.cell(startingRow+4,5).value),toNum(sheet.cell(startingRow+4,6).value))
    SPL = BinProgress(toNum(sheet.cell(startingRow+1,5).value),toNum(sheet.cell(startingRow+1,6).value))
    FCS = BinProgress(toNum(sheet.cell(startingRow+2,5).value),toNum(sheet.cell(startingRow+2,6).value))
    FT = BinProgress(toNum(sheet.cell(startingRow+3,5).value),toNum(sheet.cell(startingRow+3,6).value))
    PT = BinProgress(toNum(sheet.cell(startingRow+4,5).value),toNum(sheet.cell(startingRow+4,6).value))

    WRUObj.activities = WRUActivities(RU,JB,FB,ANT,AC,SPL,FCS,FT,PT)
    return WRUObj

WRUObjectList = []

def readWRUWorksheet():
    currRowIndex = 2
    while currRowIndex < 721:
        WRUObjectList.append(readWRUBlock(currRowIndex))
        currRowIndex += 9

def readZCaseBlock(startingRow):
    sheet = workbook["Z-CASE"]
    ZCaseObj = ZCase(toNum(sheet.cell(startingRow,1).value),str(sheet.cell(startingRow,2).value),str(sheet.cell(startingRow,3).value),None,None)

    CI = BinProgress(toNum(sheet.cell(startingRow,5).value),toNum(sheet.cell(startingRow,6).value))
    CCT = BinProgress(toNum(sheet.cell(startingRow+1,5).value),toNum(sheet.cell(startingRow+1,6).value))
    POT = BinProgress(toNum(sheet.cell(startingRow+2,5).value),toNum(sheet.cell(startingRow+2,6).value))

    ZCaseObj.activities = ZCaseActivities(CI,CCT,POT)
    return ZCaseObj

ZCaseObjectList = []

def readZCaseWorksheet():
    currRowIndex = 2
    while currRowIndex < 41:
        ZCaseObjectList.append(readZCaseBlock(currRowIndex))
        currRowIndex += 3

def readTOPBBlock(startingRow):
    sheet = workbook["TOPB"]
    TOPBObj = TOPB(toNum(sheet.cell(startingRow,1).value),str(sheet.cell(startingRow,2).value),str(sheet.cell(startingRow,3).value),None,None)

    CI = BinProgress(toNum(sheet.cell(startingRow,5).value),toNum(sheet.cell(startingRow,6).value))
    CCT = BinProgress(toNum(sheet.cell(startingRow+1,5).value),toNum(sheet.cell(startingRow+1,6).value))
    POT = BinProgress(toNum(sheet.cell(startingRow+2,5).value),toNum(sheet.cell(startingRow+2,6).value))

    TOPBObj.activities = TOPBActivities(CI,CCT,POT)
    return TOPBObj

TOPBObjectList = []

def readTOPBWorksheet():
    currRowIndex = 2
    while currRowIndex < 20:
        TOPBObjectList.append(readTOPBBlock(currRowIndex))
        currRowIndex += 3

#============================================================================ INITIALIZE ===========================================================================

def initializeObjects():
    readCMRSWorksheet()
    readAXCWorksheet()
    readSignalWorksheet()
    readSwitchWorksheet()
    readWRUWorksheet()
    readZCaseWorksheet()
    readTOPBWorksheet()
    workbook.close()

#============================================================================ USABLE FUNCTIONALITIES ===============================================================

def calcOverallProgress(objList):
    #assumes objList contains all the same objLists
    #only includes objects with valid progresses in calculation
    validCount = 0
    totalProgress = 0
    for i in range(len(objList)):
        if objList[i] != None:
            if objList[i].getProgress() != None:
                if objList[i].getProgress() != -1:
                    validCount += 1
                    totalProgress += objList[i].getProgress()

    #print(str(validCount) + " OF " + str(len(objList)) + " ENTRIES SUCCESSFULL ")
    if validCount == 0:
        return 0
    return totalProgress/validCount    

# CMRSObjectList
# AXCObjectList
# SignalObjectList
# SwitchObjectList
# WRUObjectList
# ZCaseObjectList
# TOPBObjectList
@eel.expose
def calcOverallProgressByType(equipmentType):
    match equipmentType:
        case "CMRS":
            return calcOverallProgress(CMRSObjectList)
        case "AXC":
            return calcOverallProgress(AXCObjectList)
        case "SIGNAL":
            return calcOverallProgress(SignalObjectList)
        case "SWITCH":
            return calcOverallProgress(SwitchObjectList)
        case "WRU":
            return calcOverallProgress(WRUObjectList)
        case "ZCase":
            return calcOverallProgress(ZCaseObjectList)
        case "TOPB":
            return calcOverallProgress(TOPBObjectList)
        case _:
            return -1 

def getMessengerStatsByStation(station):
    #assumes objLists are good
    #only includes
    CMSStats = {}
    messSupportProgress = 0
    messClampsProgress = 0
    messWirePullProgress = 0
    messWireTensionProgress = 0
    messCablePullProgress = 0
    messStrapProgress = 0

    total = 0
    for i in range(len(CMRSObjectList)):
        if CMRSObjectList[i] != None:
            if CMRSObjectList[i].location.upper() == station.upper() and CMRSObjectList[i].type == "mess":
                
                messSupportProgress += (CMRSObjectList[i].activities.messSupports.install / CMRSObjectList[i].activities.messSupports.total)
                messClampsProgress += (CMRSObjectList[i].activities.messClamps.install / CMRSObjectList[i].activities.messClamps.total)
                messWirePullProgress += (CMRSObjectList[i].activities.messWirePull.install / CMRSObjectList[i].activities.messWirePull.total)
                messWireTensionProgress += (CMRSObjectList[i].activities.messWireTension.install / CMRSObjectList[i].activities.messWireTension.total)
                messCablePullProgress += (CMRSObjectList[i].activities.messCablesPulled.install / CMRSObjectList[i].activities.messCablesPulled.total)
                messStrapProgress += (CMRSObjectList[i].activities.messStraps.install / CMRSObjectList[i].activities.messStraps.total)
                total += 1
    if total == 0:
        return None
    CMSStats["messSupports"] = messSupportProgress/total
    CMSStats["messClamps"] = messClampsProgress/total
    CMSStats["messWirePull"] = messWirePullProgress/total
    CMSStats["messWireTension"] = messWireTensionProgress/total
    CMSStats["messCablePull"] = messCablePullProgress/total
    CMSStats["messStrapProgress"] = messStrapProgress/total

    return CMSStats

def get15CMRSStatsByStation(station):
    #assumes objLists are good
    #only includes
    CMSStats = {}
    colClampProgress = 0
    stationBracketsProgress = 0
    groundingProgress = 0
    obsBracketProgress = 0
    cablesPulledProgress = 0
    CMRSInstallProgress = 0

    total = 0
    for i in range(len(CMRSObjectList)):
        if CMRSObjectList[i] != None:
            if CMRSObjectList[i].location.upper() == station.upper() and CMRSObjectList[i].type == "15CMRS":
                
                colClampProgress += (CMRSObjectList[i].activities.colClamp.install / CMRSObjectList[i].activities.colClamp.total)
                stationBracketsProgress += (CMRSObjectList[i].activities.stationBrackets.install / CMRSObjectList[i].activities.stationBrackets.total)
                groundingProgress += (CMRSObjectList[i].activities.grounding.install / CMRSObjectList[i].activities.grounding.total)
                obsBracketProgress += (CMRSObjectList[i].activities.obsBracket.install / CMRSObjectList[i].activities.obsBracket.total)
                cablesPulledProgress += (CMRSObjectList[i].activities.cablesPulled.install / CMRSObjectList[i].activities.cablesPulled.total)
                CMRSInstallProgress += (CMRSObjectList[i].activities.CMRSInstall15.install / CMRSObjectList[i].activities.CMRSInstall15.total)
                total += 1
    if total == 0:
        return None
    CMSStats["colClamp"] = colClampProgress/total
    CMSStats["stationBrackets"] = stationBracketsProgress/total
    CMSStats["grounding"] = groundingProgress/total
    CMSStats["obsBracket"] = obsBracketProgress/total
    CMSStats["cablesPulled"] = cablesPulledProgress/total
    CMSStats["CMRSInstall"] = CMRSInstallProgress/total


    return CMSStats

def get24CMRSStatsByStation(station):
    #assumes objLists are good
    #only includes
    CMSStats = {}
    colClampProgress = 0
    stationBracketsProgress = 0
    groundingProgress = 0
    obsBracketProgress = 0
    cablesPulledProgress = 0
    CMRSInstallProgress = 0

    total = 0
    for i in range(len(CMRSObjectList)):
        if CMRSObjectList[i] != None:
            if CMRSObjectList[i].location.upper() == station.upper() and CMRSObjectList[i].type == "24CMRS":
                
                colClampProgress += (CMRSObjectList[i].activities.colClamp.install / CMRSObjectList[i].activities.colClamp.total)
                stationBracketsProgress += (CMRSObjectList[i].activities.stationBrackets.install / CMRSObjectList[i].activities.stationBrackets.total)
                groundingProgress += (CMRSObjectList[i].activities.grounding.install / CMRSObjectList[i].activities.grounding.total)
                obsBracketProgress += (CMRSObjectList[i].activities.obsBracket.install / CMRSObjectList[i].activities.obsBracket.total)
                cablesPulledProgress += (CMRSObjectList[i].activities.cablesPulled.install / CMRSObjectList[i].activities.cablesPulled.total)
                CMRSInstallProgress += (CMRSObjectList[i].activities.CMRSInstall24.install / CMRSObjectList[i].activities.CMRSInstall24.total)
                total += 1
    if total == 0:
        return None
    CMSStats["colClamp"] = colClampProgress/total
    CMSStats["stationBrackets"] = stationBracketsProgress/total
    CMSStats["grounding"] = groundingProgress/total
    CMSStats["obsBracket"] = obsBracketProgress/total
    CMSStats["cablesPulled"] = cablesPulledProgress/total
    CMSStats["CMRSInstall"] = CMRSInstallProgress/total


    return CMSStats

def getCableSpanProgressByStation(station):
    total = 0
    validCount = 0
    for i in range(len(CMRSObjectList)):
        if CMRSObjectList[i] != None:
            if CMRSObjectList[i].location == station:
                total += CMRSObjectList[i].getProgress()
                validCount += 1
    if validCount == 0:
        return -1
    return total/validCount



def getTrayStatsByStation(station):
    #assumes objLists are good
    #only includes
    CMSStats = {}
    trayBracketProgress = 0
    coreDrillingProgress = 0
    cablePullProgress = 0
    installingTrayProgress = 0

    total = 0
    for i in range(len(CMRSObjectList)):
        if CMRSObjectList[i] != None:
            if CMRSObjectList[i].location.upper() == station.upper() and CMRSObjectList[i].type == "tray":
                
                trayBracketProgress += (CMRSObjectList[i].activities.trayBrackets.install / CMRSObjectList[i].activities.trayBrackets.total)
                coreDrillingProgress += (CMRSObjectList[i].activities.coreDrilling.install / CMRSObjectList[i].activities.coreDrilling.total)
                cablePullProgress += (CMRSObjectList[i].activities.cablePull.install / CMRSObjectList[i].activities.cablePull.total)
                installingTrayProgress += (CMRSObjectList[i].activities.installingTray.install / CMRSObjectList[i].activities.installingTray.total)
                total += 1
    
    if total == 0:
        return None
    CMSStats["trayBracket"] = trayBracketProgress/total
    CMSStats["coreDrilling"] = coreDrillingProgress/total
    CMSStats["cablePull"] = cablePullProgress/total
    CMSStats["installingTray"] = installingTrayProgress/total

    return CMSStats

def getOverallMessengerStats():
    #assumes objLists are good
    #only includes
    CMSStats = {}
    messSupportProgress = 0
    messClampsProgress = 0
    messWirePullProgress = 0
    messWireTensionProgress = 0
    messCablePullProgress = 0
    messStrapProgress = 0

    total = 0
    for i in range(len(CMRSObjectList)):
        if CMRSObjectList[i] != None:
            if CMRSObjectList[i].type == "mess":
                messSupportProgress += (CMRSObjectList[i].activities.messSupports.install / CMRSObjectList[i].activities.messSupports.total)
                messClampsProgress += (CMRSObjectList[i].activities.messClamps.install / CMRSObjectList[i].activities.messClamps.total)
                messWirePullProgress += (CMRSObjectList[i].activities.messWirePull.install / CMRSObjectList[i].activities.messWirePull.total)
                messWireTensionProgress += (CMRSObjectList[i].activities.messWireTension.install / CMRSObjectList[i].activities.messWireTension.total)
                messCablePullProgress += (CMRSObjectList[i].activities.messCablesPulled.install / CMRSObjectList[i].activities.messCablesPulled.total)
                messStrapProgress += (CMRSObjectList[i].activities.messStraps.install / CMRSObjectList[i].activities.messStraps.total)
                total += 1
    if total == 0:
        return None
    
    CMSStats["messSupports"] = messSupportProgress/total
    CMSStats["messClamps"] = messClampsProgress/total
    CMSStats["messWirePull"] = messWirePullProgress/total
    CMSStats["messWireTension"] = messWireTensionProgress/total
    CMSStats["messCablePull"] = messCablePullProgress/total
    CMSStats["messStrapProgress"] = messStrapProgress/total

    return CMSStats

def getOverall15CMRSStats():
    #assumes objLists are good
    #only includes
    CMSStats = {}
    colClampProgress = 0
    stationBracketsProgress = 0
    groundingProgress = 0
    obsBracketProgress = 0
    cablesPulledProgress = 0
    CMRSInstallProgress = 0

    total = 0
    for i in range(len(CMRSObjectList)):
        if CMRSObjectList[i] != None:
            if CMRSObjectList[i].type == "15CMRS":
                
                colClampProgress += (CMRSObjectList[i].activities.colClamp.install / CMRSObjectList[i].activities.colClamp.total)
                stationBracketsProgress += (CMRSObjectList[i].activities.stationBrackets.install / CMRSObjectList[i].activities.stationBrackets.total)
                groundingProgress += (CMRSObjectList[i].activities.grounding.install / CMRSObjectList[i].activities.grounding.total)
                obsBracketProgress += (CMRSObjectList[i].activities.obsBracket.install / CMRSObjectList[i].activities.obsBracket.total)
                cablesPulledProgress += (CMRSObjectList[i].activities.cablesPulled.install / CMRSObjectList[i].activities.cablesPulled.total)
                CMRSInstallProgress += (CMRSObjectList[i].activities.CMRSInstall15.install / CMRSObjectList[i].activities.CMRSInstall15.total)
                total += 1
    if total == 0:
        return None
    
    CMSStats["colClamp"] = colClampProgress/total
    CMSStats["stationBrackets"] = stationBracketsProgress/total
    CMSStats["grounding"] = groundingProgress/total
    CMSStats["obsBracket"] = obsBracketProgress/total
    CMSStats["cablesPulled"] = cablesPulledProgress/total
    CMSStats["CMRSInstall"] = CMRSInstallProgress/total


    return CMSStats

def getOverall24CMRSStats():
    #assumes objLists are good
    #only includes
    CMSStats = {}
    colClampProgress = 0
    stationBracketsProgress = 0
    groundingProgress = 0
    obsBracketProgress = 0
    cablesPulledProgress = 0
    CMRSInstallProgress = 0

    total = 0
    for i in range(len(CMRSObjectList)):
        if CMRSObjectList[i] != None:
            if CMRSObjectList[i].type == "24CMRS":
                colClampProgress += (CMRSObjectList[i].activities.colClamp.install / CMRSObjectList[i].activities.colClamp.total)
                stationBracketsProgress += (CMRSObjectList[i].activities.stationBrackets.install / CMRSObjectList[i].activities.stationBrackets.total)
                groundingProgress += (CMRSObjectList[i].activities.grounding.install / CMRSObjectList[i].activities.grounding.total)
                obsBracketProgress += (CMRSObjectList[i].activities.obsBracket.install / CMRSObjectList[i].activities.obsBracket.total)
                cablesPulledProgress += (CMRSObjectList[i].activities.cablesPulled.install / CMRSObjectList[i].activities.cablesPulled.total)
                CMRSInstallProgress += (CMRSObjectList[i].activities.CMRSInstall24.install / CMRSObjectList[i].activities.CMRSInstall24.total)
                total += 1
    if total == 0:
        return None
    
    CMSStats["colClamp"] = colClampProgress/total
    CMSStats["stationBrackets"] = stationBracketsProgress/total
    CMSStats["grounding"] = groundingProgress/total
    CMSStats["obsBracket"] = obsBracketProgress/total
    CMSStats["cablesPulled"] = cablesPulledProgress/total
    CMSStats["CMRSInstall"] = CMRSInstallProgress/total


    return CMSStats

def getOverallTrayStats():
    #assumes objLists are good
    #only includes
    CMSStats = {}
    trayBracketProgress = 0
    coreDrillingProgress = 0
    cablePullProgress = 0
    installingTrayProgress = 0

    total = 0
    for i in range(len(CMRSObjectList)):
        if CMRSObjectList[i] != None:
            if CMRSObjectList[i].type == "tray":
                
                trayBracketProgress += (CMRSObjectList[i].activities.trayBrackets.install / CMRSObjectList[i].activities.trayBrackets.total)
                coreDrillingProgress += (CMRSObjectList[i].activities.coreDrilling.install / CMRSObjectList[i].activities.coreDrilling.total)
                cablePullProgress += (CMRSObjectList[i].activities.cablePull.install / CMRSObjectList[i].activities.cablePull.total)
                installingTrayProgress += (CMRSObjectList[i].activities.installingTray.install / CMRSObjectList[i].activities.installingTray.total)
                total += 1
    
    if total == 0:
        return None
    CMSStats["trayBracket"] = trayBracketProgress/total
    CMSStats["coreDrilling"] = coreDrillingProgress/total
    CMSStats["cablePull"] = cablePullProgress/total
    CMSStats["installingTray"] = installingTrayProgress/total

    return CMSStats

@eel.expose
def getCMSProgressByCMSType(cmsType):
    match cmsType:
        case "mess":
            return average_dict_values(getOverallMessengerStats())
        case "15CMRS":
            return average_dict_values(getOverallMessengerStats())
        case "24CMRS":
            return average_dict_values(getOverallMessengerStats())
        case "tray":
            return average_dict_values(getOverallMessengerStats())

@eel.expose
def getCMSProgressByStationAndType(cmsType, Station):
    match cmsType:
        case "mess":
            return average_dict_values(getMessengerStatsByStation(Station))
        case "15CMRS":
            return average_dict_values(get15CMRSStatsByStation(Station))
        case "24CMRS":
            return average_dict_values(get24CMRSStatsByStation(Station))
        case "tray":
            return average_dict_values(getTrayStatsByStation(Station))

#returns dictionary
def calcProgressByStation(station, objList):
    activity_progress = {}
    count = {}
    
    for axc in objList:
        if axc.location == station or (axc.location != station and stationingToLocation(axc.stationing) == station):
            for activity_name in dir(axc.activities):
                if not activity_name.startswith("__"):
                    value = getattr(axc.activities, activity_name).progress
                    if value is not None:
                        activity_progress[activity_name] = activity_progress.get(activity_name, 0) + value
                        count[activity_name] = count.get(activity_name, 0) + 1
    
    for activity_name in activity_progress:
        activity_progress[activity_name] /= count[activity_name]
    
    return activity_progress

def calcGeneralProgressPerAttribute(objList):
    activity_progress = {}
    count = {}
    
    for axc in objList:
        for activity_name in dir(axc.activities):
            if not activity_name.startswith("__"):
                value = getattr(axc.activities, activity_name).progress
                if value is not None:
                    activity_progress[activity_name] = activity_progress.get(activity_name, 0) + value
                    count[activity_name] = count.get(activity_name, 0) + 1
    
    for activity_name in activity_progress:
        activity_progress[activity_name] /= count[activity_name]
    
    return activity_progress


#returns dictionary
@eel.expose
def calcAttributeGeneralProgressByEquipType(equipmentType):
    match equipmentType:
        case "CMRS":
            print("no")
        case "AXC":
            return calcGeneralProgressPerAttribute(AXCObjectList)
        case "SIGNAL":
            return calcGeneralProgressPerAttribute(SignalObjectList)
        case "SWITCH":
            return calcGeneralProgressPerAttribute(SwitchObjectList)
        case "WRU":
            return calcGeneralProgressPerAttribute(WRUObjectList)
        case "ZCase":
            return calcGeneralProgressPerAttribute(ZCaseObjectList)
        case "TOPB":
            return calcGeneralProgressPerAttribute(TOPBObjectList)
        case _:
            return -1

@eel.expose
def getEquipmentAttributesByStation(location, equipmentType):
    match equipmentType:
        case "CMRS":
            print("no")
        case "AXC":
            return calcProgressByStation(location,AXCObjectList)
        case "SIGNAL":
            return calcProgressByStation(location,SignalObjectList)
        case "SWITCH":
            return calcProgressByStation(location,SwitchObjectList)
        case "WRU":
            return calcProgressByStation(location,WRUObjectList)
        case "ZCase":
            return calcProgressByStation(location,ZCaseObjectList)
        case "TOPB":
            return calcProgressByStation(location,TOPBObjectList)
        case _:
            return -1


@eel.expose
def calcProgressByLocation(location, equipmentType):
    match equipmentType:
        case "CMRS":
            return getCableSpanProgressByStation(location)
        case "AXC":
            return average_dict_values(calcProgressByStation(location,AXCObjectList))
        case "SIGNAL":
            return average_dict_values(calcProgressByStation(location,SignalObjectList))
        case "SWITCH":
            return average_dict_values(calcProgressByStation(location,SwitchObjectList))
        case "WRU":
            return average_dict_values(calcProgressByStation(location,WRUObjectList))
        case "ZCase":
            return average_dict_values(calcProgressByStation(location,ZCaseObjectList))
        case "TOPB":
            return average_dict_values(calcProgressByStation(location,TOPBObjectList))
        case "GENERAL":
            total = 0
            numAttributes = 0
            if getCableSpanProgressByStation(location) != None:
                if getCableSpanProgressByStation(location) > 0:
                    total += getCableSpanProgressByStation(location)
                    numAttributes += 1
            if average_dict_values(calcProgressByStation(location,AXCObjectList)) > 0:
                total += average_dict_values(calcProgressByStation(location,AXCObjectList))
                numAttributes += 1
            if average_dict_values(calcProgressByStation(location,SignalObjectList)) > 0:
                total += average_dict_values(calcProgressByStation(location,SignalObjectList))
                numAttributes += 1
            if average_dict_values(calcProgressByStation(location,SwitchObjectList)) > 0:
                total += average_dict_values(calcProgressByStation(location,SwitchObjectList))
                numAttributes += 1
            if average_dict_values(calcProgressByStation(location,WRUObjectList)) > 0:
                total += average_dict_values(calcProgressByStation(location,WRUObjectList))
                numAttributes += 1
            if average_dict_values(calcProgressByStation(location,ZCaseObjectList)) > 0:
                total += average_dict_values(calcProgressByStation(location,ZCaseObjectList))
                numAttributes += 1
            if average_dict_values(calcProgressByStation(location,TOPBObjectList)) > 0:
                total += average_dict_values(calcProgressByStation(location,TOPBObjectList))
                numAttributes += 1
            if numAttributes == 0:
                return -1
            return total/numAttributes


        case _:
            return -1 

def average_dict_values(d):
    if not d:
        return -1
    return sum(d.values()) / len(d)

def stationingToLocation(stationing):
    if stationing == None:
        return None
    stationing = int(stationing)
    if stationing > 55696 and stationing < 56356:
        return "CHURCH"
    elif stationing < 58488:
        return "CHU-FOR"
    elif stationing < 59190:
        return "FORT HAMILTON"
    elif stationing < 62570:
        return "FOR-15S"
    elif stationing < 63230:
        return "15TH STREET PROSPECT PARK"
    elif stationing < 65396:
        return "15S-7AV"
    elif stationing < 66056:
        return "7TH AV"
    elif stationing < 68383:
        return "7AV-4TH"
    elif stationing < 69091:
        return "4TH AVE"
    elif stationing < 70511:
        return "4TH-SMI"
    elif stationing < 71232:
        return "SMITH & 9TH STREET"
    elif stationing < 73289:
        return "SMI-CAR"
    elif stationing < 73899:
        return "CARROLL"
    elif stationing < 75682:
        return "CAR-BER"
    elif stationing < 76342:   
        return "BERGEN"
    elif stationing < 95182: 
        return None               #STATIOINGS ARE INACCURATE HERE, INCLUDE WARNINGS
    elif stationing < 95842:
        return "CLINTON-WASHINGTON"
    elif stationing < 97082:
        return "CLI-CLA"
    elif stationing < 97742:
        return "CLASSON"
    elif stationing < 98922:
        return "CLA-BED"
    elif stationing < 99582:
        return "BEDFORD NOSTRAND"
    elif stationing < 101747:
        return "BED-MYR"
    elif stationing < 102452:
        return "MYRTLE AVE"
    elif stationing < 103942:
        return "MYR-FLU"
    elif stationing < 104620:
        return "FLUSHING"
    elif stationing < 106052:
        return "FLU-BRO"
    elif stationing < 106712:
        return "BROADWAY"
    elif stationing < 113245:
        return "BRO-MET"
    elif stationing < 113950:
        return None               #STATIOINGS ARE INACCURATE HERE, INCLUDE WARNINGS
    elif stationing < 115700:
        return "NAS-GRE"
    elif stationing < 116432:
        return "GREENPOINT"
    elif stationing < 121133:
        return "GRE-21S"
    elif stationing < 121792:
        return "21 STREET"
    elif stationing < 122923:
        return "21S-COU"
    elif stationing < 123666:
        return "COURT SQ"
    elif stationing < 125065:
        return "COU-QUE"
    elif stationing < 125727:
        return "QUEENS PLAZA"
#============================================================================ CODE EXECUTION START =================================================================
initializeObjects()

#============================================================================= INITIALIZE PYTHON EEL WINDOW ======================================
if __name__ == "__main__":   
    if 'win' in sys.platform:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)

web_path = current_directory + "\\web"
eel.init(web_path)
try:
    eel.start("index.html",size=(960*1.5,540*1.5))
except (SystemExit, MemoryError, KeyboardInterrupt):
    os.system("taskkill /F /IM python.exe /T")

