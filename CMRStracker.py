from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import os

#============================= REMOVE DURING DEBUGGING ==========================================
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
workbook = load_workbook(filename="C:\\Users\\mdas\\Documents\\CMRSTracker\\dashboard.xlsx",  data_only=True)
#================================================================================================
#========================================================================================================================================================================
#====================================================== CLASS DEFINITIONS, SEE GITHUB REPO FOR DETAILS ==================================================================
#========================================================================================================================================================================

#======================================================= GENERAL USE CLASSES/SUPERCLASSES =============================================================================

class BinProgress:
    def __init__(self, date, progress):
        self.date = date
        self.progress = progress

class Equipment:
    def __init__(self,type,stationing,track,location,activities,notes):
        self.type = type
        self.stationing = stationing
        self.track = track
        self.location = location
        self.activities = activities
        self.notes = notes

#====================================================== CMRS CLASSES ==============================================================================================
class CableSpan:
    def __init__(self,start,end,track,location,type,activities,notes):
        self.start = start
        self.end = end
        self.track = track
        self.location = location
        self.type = type
        self.activities = activities
        self.notes = notes

    def __str__(self):
        return str(self.type).upper() + " " + str(self.start) + " to " + str(self.end)
    
class MessActivities:
    def __init__(self, messSupports, messClamps, messWirePull, messWireTension, messCablesPulled,messStraps):
        self.messSupports = messSupports
        self.messClamps = messClamps
        self.messWirePull = messWirePull
        self.messWireTension = messWireTension
        self.messCablesPulled = messCablesPulled
        self.messStraps = messStraps
        
class CMRSActivities:
    def __init__(self, colClamp, stationBrackets, grounding, obsBracket, cablesPulled):
        self.colClamp = colClamp
        self.stationBrackets = stationBrackets
        self.grounding = grounding
        self.obsBracket = obsBracket
        self.cablesPulled = cablesPulled

class CMRS15Activities(CMRSActivities):
    def __init__(self, CMRSInstall15, colClamp, stationBrackets, grounding, obsBracket, cablesPulled):
        super().__init__(colClamp, stationBrackets, grounding, obsBracket, cablesPulled)
        self.CMRSInstall15 = CMRSInstall15

class CMRS24Activities(CMRSActivities):
    def __init__(self, CMRSInstall24, colClamp, stationBrackets, grounding, obsBracket, cablesPulled):
        super().__init__(colClamp, stationBrackets, grounding, obsBracket, cablesPulled)
        self.CMRSInstall24 = CMRSInstall24

class TrayActivities:
    def __init__(self, trayBrackets, installingTray, coreDrilling, cablePull):
        self.trayBrackets = trayBrackets
        self.installingTray = installingTray
        self.coreDrilling = coreDrilling
        self.cablePull = cablePull

class progress:
    def __init__(self, total, install):
        self.total = total
        self.install = install

    def getProgress(self):
        if self.total == None or self.install == None:
            return None
        if self.total > 0:
            return self.install/self.total * 100
        else:
            raise ValueError("the \"total\" of this activity is either 0 or not defined")

#========================================================================== AXLE COUNTER ============================================================================

class AxleCounterActivities:
    def __init__(self,ACInstall, JBInstall, LCInstall, ECInstall, preOpTesting):
        self.ACInstall = ACInstall
        self.JBInstall = JBInstall
        self.LCInstall = LCInstall
        self.ECInstall = ECInstall
        self.preOpTesting = preOpTesting

#======================================================================== SIGNALS ===================================================================================

class SignalActivities:
    def __init__(self, sigInstall, JBInstall, SMInstall, LCInstall, breakdownTesting, preOpTesting):
        self.sigInstall = sigInstall
        self.JBInstall = JBInstall
        self.SMInstall = SMInstall
        self.LCInstall = LCInstall
        self.breakdownTesting = breakdownTesting
        self.preOpTesting = preOpTesting

# ==================================================================== SWITCH ======================================================================================

class SwitchActivities:
    def __init__(self, switchInstall, JBInstall, SCInstall, LCInstall, breakdownTesting, preOpTesting):
        self.switchInstall = switchInstall
        self.JBInstall = JBInstall
        self.SCInstall = SCInstall
        self.LCInstall = LCInstall
        self.breakdownTesting = breakdownTesting
        self.preOpTesting = preOpTesting

# ==================================================================== WRU ========================================================================================

class WRUActivities:
    def __init__(self, RUInstall, JBInstall, FBInstall, antennaInstall, antCableInstall, splitterInstall, FCSplice, FTesting, PTesting):
        self.RUInstall = RUInstall
        self.JBInstall = JBInstall
        self.FBInstall = FBInstall
        self.antennaInstall = antennaInstall
        self.antCableInstall = antCableInstall
        self.splitterInstall = splitterInstall
        self.FCSplice = FCSplice
        self.FTesting = FTesting
        self.PTesting = PTesting

#============================================================ END CLASS DEFINITIONS ============================================================================

def toNum(intVal):
    if intVal == None:
        return None
    try:
        return int(intVal)
    except ValueError:
        return None

def readCMRSBlock(startingRow):
    sheet = workbook["CMRS"]
    match sheet.cell(startingRow,5).value:
        case "Steel Mess Supports":
            MessengerSpan = CableSpan(toNum(sheet.cell(startingRow,1).value),toNum(sheet.cell(startingRow,2).value),str(sheet.cell(startingRow,3).value),str(sheet.cell(startingRow,4).value),"mess",None,None)
            
            sMSupp = progress(toNum(sheet.cell(startingRow,6).value),toNum(sheet.cell(startingRow,7).value))
            mClamps = progress(toNum(sheet.cell(startingRow+1,6).value),toNum(sheet.cell(startingRow+1,7).value))
            mWPull = progress(toNum(sheet.cell(startingRow+2,6).value),toNum(sheet.cell(startingRow+2,7).value))
            mWTension = progress(toNum(sheet.cell(startingRow+3,6).value),toNum(sheet.cell(startingRow+3,7).value))
            nCPull = progress(toNum(sheet.cell(startingRow+4,6).value),toNum(sheet.cell(startingRow+4,7).value))
            mStraps = progress(toNum(sheet.cell(startingRow+5,6).value),toNum(sheet.cell(startingRow+5,7).value))
            
            MessengerSpan.activities = MessActivities(sMSupp,mClamps,mWPull,mWTension,nCPull,mStraps)
            return MessengerSpan
        
        case "15\" CMRS Install":
            CMRS15Span = CableSpan(toNum(sheet.cell(startingRow,1).value),toNum(sheet.cell(startingRow,2).value),str(sheet.cell(startingRow,3).value),str(sheet.cell(startingRow,4).value),"15CMRS",None,None)
            
            cInstall = progress(toNum(sheet.cell(startingRow,6).value),toNum(sheet.cell(startingRow,7).value))
            cClamp = progress(toNum(sheet.cell(startingRow+1,6).value),toNum(sheet.cell(startingRow+1,7).value))
            sBracket = progress(toNum(sheet.cell(startingRow+2,6).value),toNum(sheet.cell(startingRow+2,7).value))
            grounding = progress(toNum(sheet.cell(startingRow+3,6).value),toNum(sheet.cell(startingRow+3,7).value))
            oBracket = progress(toNum(sheet.cell(startingRow+4,6).value),toNum(sheet.cell(startingRow+4,7).value))
            nCPull = progress(toNum(sheet.cell(startingRow+5,6).value),toNum(sheet.cell(startingRow+5,7).value))

            CMRS15Span.activities = CMRS15Activities(cInstall,cClamp,sBracket,grounding,oBracket,nCPull)
            return CMRS15Span
        
        case "24\" CMRS Install":
            CMRS24Span = CableSpan(toNum(sheet.cell(startingRow,1).value),toNum(sheet.cell(startingRow,2).value),str(sheet.cell(startingRow,3).value),str(sheet.cell(startingRow,4).value),"24CMRS",None,None)
            
            cInstall = progress(toNum(sheet.cell(startingRow,6).value),toNum(sheet.cell(startingRow,7).value))
            cClamp = progress(toNum(sheet.cell(startingRow+1,6).value),toNum(sheet.cell(startingRow+1,7).value))
            sBracket = progress(toNum(sheet.cell(startingRow+2,6).value),toNum(sheet.cell(startingRow+2,7).value))
            grounding = progress(toNum(sheet.cell(startingRow+3,6).value),toNum(sheet.cell(startingRow+3,7).value))
            oBracket = progress(toNum(sheet.cell(startingRow+4,6).value),toNum(sheet.cell(startingRow+4,7).value))
            nCPull = progress(toNum(sheet.cell(startingRow+5,6).value),toNum(sheet.cell(startingRow+5,7).value))

            CMRS24Span.activities = CMRS24Activities(cInstall,cClamp,sBracket,grounding,oBracket,nCPull)
            return CMRS24Span
        case "Cable Tray Brackets":
            TraySpan = CableSpan(toNum(sheet.cell(startingRow,1).value),toNum(sheet.cell(startingRow,2).value),str(sheet.cell(startingRow,3).value),str(sheet.cell(startingRow,4).value),"tray",None,None)

            ctBrackets = progress(toNum(sheet.cell(startingRow,6).value),toNum(sheet.cell(startingRow,7).value))
            iTray = progress(toNum(sheet.cell(startingRow+1,6).value),toNum(sheet.cell(startingRow+1,7).value))
            cDrilling = progress(toNum(sheet.cell(startingRow+2,6).value),toNum(sheet.cell(startingRow+2,7).value))
            cPull = progress(toNum(sheet.cell(startingRow+3,6).value),toNum(sheet.cell(startingRow+3,7).value))

            TraySpan.activities = TrayActivities(ctBrackets,iTray,cDrilling,cPull)
            return TraySpan
        case "Crossover (Arch Bar Installation)":
            return None
        case _:
            err_message = "ERR | " + str(sheet.cell(startingRow,5).value) + "is NOT recognized as a cablespan type"
            print(err_message)


CMRSObjectList = []

def readCMRSWorksheet():
    currRowIndex = 3
    while currRowIndex <= 1215:
        block1 = readCMRSBlock(currRowIndex)
        CMRSObjectList.append(block1)
        if block1 == None:
            currRowIndex += 1
        else:
            match block1.type:
                case "mess":
                    currRowIndex += 6
                case "15CMRS":
                    currRowIndex += 6
                case "24CMRS":
                    currRowIndex += 6
                case "tray":
                    currRowIndex += 4

AXCObjectList = []

def readAXCBlock(startingRow):
    sheet = workbook["AXLE COUNTER"]
    equipObj = Equipment("AXC",toNum(sheet.cell(startingRow,1).value),toNum(sheet.cell(startingRow,2).value),str(sheet.cell(startingRow,3).value),None,None)

    AC = BinProgress(toNum(sheet.cell(startingRow,5).value),toNum(sheet.cell(startingRow,6).value))
    JB = BinProgress(toNum(sheet.cell(startingRow+1,5).value),toNum(sheet.cell(startingRow+1,6).value))
    LC = BinProgress(toNum(sheet.cell(startingRow+2,5).value),toNum(sheet.cell(startingRow+2,6).value))
    EC = BinProgress(toNum(sheet.cell(startingRow+3,5).value),toNum(sheet.cell(startingRow+3,6).value))
    POT = BinProgress(toNum(sheet.cell(startingRow+4,5).value),toNum(sheet.cell(startingRow+4,6).value))

    equipObj.activities = AxleCounterActivities(AC,JB,LC,EC,POT)
    return equipObj

def readAXCBlock(startingRow):
    sheet = workbook["AXLE COUNTER"]
    equipObj = Equipment("AXC",toNum(sheet.cell(startingRow,1).value),toNum(sheet.cell(startingRow,2).value),str(sheet.cell(startingRow,3).value),None,None)

    AC = BinProgress(toNum(sheet.cell(startingRow,5).value),toNum(sheet.cell(startingRow,6).value))
    JB = BinProgress(toNum(sheet.cell(startingRow+1,5).value),toNum(sheet.cell(startingRow+1,6).value))
    LC = BinProgress(toNum(sheet.cell(startingRow+2,5).value),toNum(sheet.cell(startingRow+2,6).value))
    EC = BinProgress(toNum(sheet.cell(startingRow+3,5).value),toNum(sheet.cell(startingRow+3,6).value))
    POT = BinProgress(toNum(sheet.cell(startingRow+4,5).value),toNum(sheet.cell(startingRow+4,6).value))

    equipObj.activities = AxleCounterActivities(AC,JB,LC,EC,POT)
    return equipObj
#def readAXCWorksheet():
#    currRowIndex = 2
#    while currRowIndex <= 397:


#============================================================================ CODE EXECUTION START =================================================================

#readCMRSWorksheet()
